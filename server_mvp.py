from flask import Flask, request, jsonify
from flask_cors import CORS
from pymongo import MongoClient
from openpyxl import load_workbook
import os
import gc
import threading
import time
from datetime import datetime

app = Flask(__name__)
CORS(app)

def keep_alive_ping():
    """Ping MongoDB every 6 hours to prevent Atlas free tier from pausing."""
    while True:
        time.sleep(6 * 60 * 60)  # 6 hours
        try:
            database = get_db()
            if database is not None:
                database.client.admin.command('ping')
                print("Keep-alive ping sent to MongoDB")
        except Exception as e:
            print(f"Keep-alive ping failed: {e}")

# Start keep-alive thread
ping_thread = threading.Thread(target=keep_alive_ping, daemon=True)
ping_thread.start()

# MongoDB configuration
MONGO_URI = os.getenv("MONGO_URI")
mongo_client = None
db = None
COLLECTION_NAME = 'danalog_catalog'

def get_db():
    """Get database connection, initialize if needed. Reconnects on stale connections."""
    global mongo_client, db
    if not MONGO_URI:
        return None
    if db is not None:
        # Verify the connection is still alive
        try:
            mongo_client.admin.command('ping')
            return db
        except Exception:
            # Connection is stale, reset and reconnect
            print("MongoDB connection stale, reconnecting...")
            mongo_client = None
            db = None
    try:
        mongo_client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=30000)
        db = mongo_client['ayal_kore']
        mongo_client.admin.command('ping')
        print("MongoDB connected successfully")
        return db
    except Exception as e:
        print(f"MongoDB connection failed: {e}")
        mongo_client = None
        db = None
        return None

def init_db():
    """Create indexes if they don't exist"""
    database = get_db()
    if database is None:
        return

    collection = database[COLLECTION_NAME]

    # Create unique index on דאנאקוד
    collection.create_index("דאנאקוד", unique=True, sparse=True)

    # Create indexes for faster searching
    collection.create_index("שם")
    collection.create_index("מחבר")
    collection.create_index("נושא")
    collection.create_index("ברקוד")

@app.route('/')
def index():
    """Health check endpoint"""
    database = get_db()
    return jsonify({
        'status': 'running',
        'message': 'Ayal Kore Catalog Server',
        'database_connected': database is not None
    })

@app.route('/upload', methods=['POST'])
def upload_danalog():
    """Upload Danalog Excel file and import into MongoDB (optimized for speed)"""
    database = get_db()
    if database is None:
        return jsonify({'error': 'Database not available. Please check MongoDB connection.'}), 500

    try:
        if 'file' not in request.files:
            return jsonify({'error': 'לא נבחר קובץ'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'לא נבחר קובץ'}), 400

        # Initialize database indexes
        init_db()
        collection = database[COLLECTION_NAME]

        # Fetch all existing דאנאקוד codes in ONE query (fast lookup)
        existing_codes = set(
            doc['דאנאקוד'] for doc in collection.find({}, {'דאנאקוד': 1, '_id': 0})
            if doc.get('דאנאקוד')
        )

        # Use openpyxl with read_only mode for memory efficiency
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active

        docs_to_insert = []
        skipped_count = 0
        headers = None

        for row_num, row in enumerate(ws.iter_rows(values_only=True)):
            # First row is headers
            if row_num == 0:
                headers = list(row)
                continue

            # Skip empty rows
            if not any(row):
                continue

            # Build document from row
            doc = {}
            danalog_code = None
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    col_name = headers[i]
                    doc[col_name] = value if value is not None else None
                    if col_name == 'דאנאקוד':
                        danalog_code = value

            # Check for duplicate using in-memory set (instant)
            if danalog_code and danalog_code in existing_codes:
                skipped_count += 1
                continue

            # Add timestamp and queue for batch insert
            doc['created_at'] = datetime.utcnow()
            docs_to_insert.append(doc)
            if danalog_code:
                existing_codes.add(danalog_code)  # Prevent duplicates within file

        # Close workbook and free memory
        wb.close()

        # Batch insert all documents at once
        added_count = 0
        if docs_to_insert:
            try:
                result = collection.insert_many(docs_to_insert, ordered=False)
                added_count = len(result.inserted_ids)
            except Exception as e:
                # Handle partial failures (some duplicates)
                if hasattr(e, 'details'):
                    added_count = e.details.get('nInserted', 0)
                    skipped_count += len(docs_to_insert) - added_count

        gc.collect()
        total_count = collection.count_documents({})

        return jsonify({
            'success': True,
            'added': added_count,
            'skipped': skipped_count,
            'total': total_count
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/search', methods=['GET'])
def search_catalog():
    """Search catalog by column and text"""
    database = get_db()
    if database is None:
        return jsonify({'error': 'Database not available. Please check MongoDB connection.'}), 500

    try:
        column = request.args.get('column')
        search_text = request.args.get('text')

        if not column or not search_text:
            return jsonify({'error': 'חסרים פרמטרים לחיפוש'}), 400

        collection = database[COLLECTION_NAME]

        # Check if database has data
        if collection.count_documents({}) == 0:
            return jsonify({'error': 'מסד הנתונים ריק. אנא העלה קטלוג תחילה'}), 400

        # Build search query - use regex for partial matching
        query = {column: {"$regex": search_text, "$options": "i"}}

        # Projection to return only important fields
        projection = {
            "_id": 0,
            "ID": 1,
            "דאנאקוד": 1,
            "שם": 1,
            "ת.מחלקה": 1,
            "מחיר": 1,
            "מחבר": 1,
            "נושא": 1,
            "ברקוד": 1,
            "ת.פתיחה": 1,
            "ת.עדכון": 1,
            "ת.מה.ראשונה": 1
        }

        results = list(collection.find(query, projection))

        return jsonify({
            'success': True,
            'results': results,
            'count': len(results)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/stats', methods=['GET'])
def get_stats():
    """Get database statistics"""
    database = get_db()
    if database is None:
        return jsonify({
            'database_exists': False,
            'total_books': 0,
            'message': 'Database not available. Please check MongoDB connection.'
        })

    try:
        collection = database[COLLECTION_NAME]
        total_count = collection.count_documents({})

        return jsonify({
            'database_exists': True,
            'total_books': total_count
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    init_db()
    print("Starting Ayal Kore Catalog Server...")
    app.run(debug=True, host='0.0.0.0', port=5000)
